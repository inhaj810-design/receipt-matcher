import streamlit as st
import pandas as pd
import google.generativeai as genai
from PIL import Image
import json
import math
import io
import zipfile
import time
import re
from datetime import datetime
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="영수증 HCP 매칭", page_icon="🧾", layout="wide")

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Noto+Sans+KR:wght@300;400;500;700&display=swap');
html, body, [class*="css"] { font-family: 'Noto Sans KR', sans-serif; }
.stApp { background: #f8f9fc; }
.hero {
    background: linear-gradient(135deg, #1a1a2e 0%, #16213e 50%, #0f3460 100%);
    border-radius: 16px; padding: 40px 48px; margin-bottom: 32px;
    color: white; position: relative; overflow: hidden;
}
.hero h1 { font-size: 2rem; font-weight: 700; margin: 0 0 8px 0; }
.hero p  { font-size: 0.95rem; opacity: 0.7; margin: 0; }
.step-card {
    background: white; border-radius: 12px; padding: 24px; margin-bottom: 16px;
    border: 1px solid #e8ecf0; box-shadow: 0 2px 8px rgba(0,0,0,0.04);
}
.step-num {
    display: inline-block; background: #0f3460; color: white;
    width: 28px; height: 28px; border-radius: 50%; text-align: center;
    line-height: 28px; font-size: 13px; font-weight: 700; margin-right: 10px;
}
.step-title { font-size: 1rem; font-weight: 600; color: #1a1a2e; }
</style>
""", unsafe_allow_html=True)

st.markdown("""
<div class="hero">
    <h1>🧾 영수증 HCP 매칭 시스템</h1>
    <p>영수증 사진과 방문기록 엑셀을 업로드하면 자동으로 HCP를 매칭해드립니다</p>
</div>
""", unsafe_allow_html=True)

# ─── STEP 1: API KEY ────────────────────────────────────
st.markdown("""<div class="step-card">
<span class="step-num">1</span><span class="step-title">Gemini API 키 입력</span>
</div>""", unsafe_allow_html=True)
api_key = st.text_input("Gemini API Key", type="password", placeholder="AIza...")
if api_key:
    genai.configure(api_key=api_key)
    st.success("✅ API 키 연결됨")

# ─── STEP 2: 규정 설정 ─────────────────────────────────
st.markdown("""<div class="step-card">
<span class="step-num">2</span><span class="step-title">매칭 규정 설정</span>
</div>""", unsafe_allow_html=True)
col1, col2 = st.columns(2)
with col1:
    min_days = st.number_input("동일 HCP 최소 간격 (일)", min_value=1, max_value=30, value=7)
with col2:
    max_per_month = st.number_input("월 최대 배정 횟수", min_value=1, max_value=10, value=4)

# ─── STEP 3: 파일 업로드 ────────────────────────────────
st.markdown("""<div class="step-card">
<span class="step-num">3</span><span class="step-title">파일 업로드</span>
</div>""", unsafe_allow_html=True)
col1, col2 = st.columns(2)
with col1:
    receipt_files = st.file_uploader(
        "영수증 사진 (여러 장 가능)",
        type=["png", "jpg", "jpeg"],
        accept_multiple_files=True
    )
    if receipt_files:
        st.info(f"📸 {len(receipt_files)}장 업로드됨")
with col2:
    excel_file = st.file_uploader("방문기록 엑셀", type=["xlsx", "xls"])
    if excel_file:
        st.info(f"📊 {excel_file.name} 업로드됨")

# ─── 품목별 진료과 매핑 ──────────────────────────────────
PRODUCT_DEPT = {
    '오구멘틴':      ['이비인후과', '소아청소년과', '외과', '감염내과', '호흡기내과'],
    '클래리시드 정주': ['감염내과', '호흡기내과', '소아청소년과'],
    '호의주':        ['소화기내과'],
    '호이판':        ['소화기내과'],
    '원알파':        ['신장내과', '내분비내과', '유방갑상선외과', '갑상선외과', '내분비외과'],
    '이솦틴':        ['심장내과'],
    '카라듀오':      ['소화기내과'],
    '피에젯타':      ['순환기내과', '심장내과', '내분비내과'],
}

def get_product_for_dept(dept):
    """진료과로 가능한 품목 리스트 반환"""
    products = []
    for prod, depts in PRODUCT_DEPT.items():
        if any(d in dept for d in depts):
            products.append(prod)
    return products

def select_best_product(candidates, n):
    """
    후보 HCP 목록에서 n명을 같은 품목으로 묶을 수 있는 최적 품목 선택
    - 같은 품목 가능한 HCP가 n명 이상이면 그 품목 선택
    - 없으면 가장 많은 HCP를 묶을 수 있는 품목 선택
    반환: (품목명, 해당 품목 가능한 후보 리스트)
    """
    product_candidates = defaultdict(list)
    for hosp, v in candidates:
        dept = v.get('진료과', '')
        prods = get_product_for_dept(dept)
        for prod in prods:
            product_candidates[prod].append((hosp, v))

    if not product_candidates:
        return None, []

    # n명 이상 묶을 수 있는 품목 중 등급 높은 HCP 많은 순
    best_product = None
    best_list = []
    best_count = 0

    for prod, cands in product_candidates.items():
        cnt = len(cands)
        if cnt >= n:
            # n명 충족 → 바로 선택 (등급순 정렬 후 앞에서 n명)
            if best_product is None or cnt > best_count:
                best_product = prod
                best_list = cands
                best_count = cnt

    if best_product:
        return best_product, best_list

    # n명 충족 못하면 가장 많은 품목
    for prod, cands in sorted(product_candidates.items(), key=lambda x: -len(x[1])):
        if len(cands) > best_count:
            best_product = prod
            best_list = cands
            best_count = len(cands)

    return best_product, best_list

# ─── 영수증 분석 함수 ────────────────────────────────────
def analyze_receipt_with_gemini(image_file):
    model = genai.GenerativeModel('gemini-2.5-flash-lite')
    img = Image.open(image_file)
    prompt = """이 영수증 사진에서 다음 정보를 추출해서 JSON으로만 응답해줘. 다른 텍스트 없이 JSON만.
{
  "shop_name": "가맹점명 (영수증에 적힌 그대로)",
  "address": "주소 (있으면, 없으면 빈 문자열)",
  "date": "YYYY-MM-DD",
  "time": "HH:MM",
  "amount": 총결제금액(숫자만),
  "approval_number": "승인번호",
  "items": [
    {"name": "품목명", "price": 단가(숫자), "qty": 수량(숫자)}
  ]
}
items는 영수증에 있는 품목들을 모두 추출해줘. price는 1개당 단가야."""

    for attempt in range(10):
        try:
            response = model.generate_content([prompt, img])
            text = response.text.strip()
            if text.startswith("```"):
                text = text.split("```")[1]
                if text.startswith("json"):
                    text = text[4:]
            return json.loads(text.strip())
        except Exception as e:
            err = str(e)
            if "429" in err:
                wait = 70
                m = re.search(r"seconds: (\d+)", err)
                if m:
                    wait = int(m.group(1)) + 10
                st.warning(f"  ⏳ API 한도 초과 → {wait}초 대기 후 재시도 ({attempt+1}/10)...")
                time.sleep(wait)
            else:
                raise
    raise Exception("최대 재시도 횟수(10회) 초과")

def has_item_over_10000(info):
    for item in info.get('items', []):
        try:
            if float(item.get('price', 0)) > 10000:
                return True
        except:
            pass
    return False

def load_visit_records(f):
    df = pd.read_excel(f, sheet_name=0, header=None)
    visits = []
    for i in range(2, len(df)):
        row = df.iloc[i]
        if pd.notna(row[5]) and pd.notna(row[10]) and pd.notna(row[11]):
            visit_date = str(row[5])[:10]
            hospital = str(row[10])
            hcp = str(row[11])
            dept  = str(row[13]) if pd.notna(row[13]) else ''
            grade = str(row[12]) if pd.notna(row[12]) else ''
            if hospital not in ['출근', '퇴근', '']:
                visits.append({'방문일자': visit_date, '거래처': hospital,
                                'HCP': hcp, '진료과': dept, '등급': grade})
    return visits

def guess_hospital(shop_name, address, date, vdh):
    keywords = {
        '안동성소': ['안동성소병원'], '안동병원': ['의료법인안동병원'],
        '가톨릭대': ['대구가톨릭대학교병원'], '가톨릭병원': ['대구가톨릭대학교병원'],
        '계명대': ['계명대학교동산병원'], '동산병원': ['계명대학교동산병원'],
        '칠곡경북대': ['칠곡경북대학교병원'], '경북대병원': ['칠곡경북대학교병원'],
        '구미': ['순천향대학교 부속 구미병원'], '순천향': ['순천향대학교 부속 구미병원'],
        '곽병원': ['곽병원'],
    }
    text = (shop_name or '') + (address or '')
    for kw, h in keywords.items():
        if kw in text:
            return h
    if address:
        if '안동' in address:
            r = [h for h in vdh.get(date, []) if '안동' in h]
            if r: return r
        if '구미' in address:
            return ['순천향대학교 부속 구미병원']
        if '대구' in address:
            r = [h for h in vdh.get(date, []) if any(k in h for k in ['대구','계명','칠곡','곽'])]
            if r: return r
    return vdh.get(date, [])

def calc_persons(amount):
    return math.ceil(amount / 10000)

# ─── HCP 매칭 ────────────────────────────────────────────
def match_hcp(receipts, visits, min_days, max_per_month):
    grade_order = {'S':0,'A1':1,'A2':2,'B1':3,'B2':4,'B3':5,'C1':6,'C2':7,'D1':8,'D2':9}
    def gk(v): return grade_order.get(v['등급'], 99)

    vbdh = defaultdict(list)
    for v in visits:
        vbdh[(v['방문일자'], v['거래처'])].append(v)
    vdh = defaultdict(list)
    for v in visits:
        if v['거래처'] not in vdh[v['방문일자']]:
            vdh[v['방문일자']].append(v['거래처'])

    hcp_dates = defaultdict(list)

    def can_assign(hcp, ds):
        d = datetime.strptime(ds, '%Y-%m-%d')
        for pd_ in hcp_dates[hcp]:
            if abs((d - pd_).days) < min_days: return False
        if sum(1 for pd_ in hcp_dates[hcp] if pd_.month==d.month and pd_.year==d.year) >= max_per_month:
            return False
        return True

    results = []
    for r in receipts:
        date, n = r['date'], r['persons']

        # 품목 단가 1만원 초과 → 서명 필요
        if r.get('need_sign'):
            results.append({
                'No': r['no'], '승인일자': date, '시간': r['time'],
                '병원': '-', '품목': '-', 'HCP': '참석자 명단 서명 필요',
                '진료과': '', '가맹점': r['shop_name'], '주소': r.get('address',''),
                '승인금액': r['amount'], '선정인원': '-',
                '비고': '⚠ 품목 단가 1만원 초과 (별도 지출보고서)'
            })
            continue

        hospitals = r.get('hospitals') or guess_hospital(
            r.get('shop_name',''), r.get('address',''), date, vdh)

        # 전체 후보 수집 (7일/월4회 가능한 HCP만)
        all_candidates = []
        for hosp in (hospitals or []):
            for c in vbdh.get((date, hosp), []):
                if can_assign(c['HCP'], date):
                    all_candidates.append((hosp, c))
        all_candidates.sort(key=lambda x: gk(x[1]))

        # 최적 품목 선택 (n명을 같은 품목으로 묶기)
        best_product, prod_candidates = select_best_product(all_candidates, n)

        # 품목 후보에서 등급순 n명 선택
        prod_candidates.sort(key=lambda x: gk(x[1]))
        selected = []
        for hosp, v in prod_candidates:
            if len(selected) >= n: break
            selected.append((hosp, v))
            hcp_dates[v['HCP']].append(datetime.strptime(date, '%Y-%m-%d'))

        shortage = n - len(selected)
        warn = f'⚠ {n}명 필요 / {len(selected)}명 배정 (품목 해당 HCP 부족)' if shortage > 0 else ''
        if not best_product:
            warn = '⚠ 매칭 가능한 품목/HCP 없음'

        if selected:
            for idx, (hosp, v) in enumerate(selected):
                results.append({
                    'No': r['no'], '승인일자': date, '시간': r['time'],
                    '병원': hosp, '품목': best_product or '-',
                    'HCP': v['HCP'], '진료과': v['진료과'],
                    '가맹점': r['shop_name'], '주소': r.get('address',''),
                    '승인금액': r['amount'], '선정인원': n,
                    '비고': warn if idx == 0 else ''
                })
        else:
            results.append({
                'No': r['no'], '승인일자': date, '시간': r['time'],
                '병원': ', '.join(hospitals or ['미확인']),
                '품목': best_product or '-', 'HCP': '(매칭 없음)',
                '진료과': '', '가맹점': r['shop_name'], '주소': r.get('address',''),
                '승인금액': r['amount'], '선정인원': n,
                '비고': warn or '⚠ HCP 없음'
            })
    return results

# ─── 엑셀 생성 ───────────────────────────────────────────
def make_excel(results):
    wb = Workbook(); ws = wb.active; ws.title = '매칭결과'
    hf    = Font(name='Malgun Gothic', bold=True, color='FFFFFF', size=10)
    hfill = PatternFill('solid', start_color='1a1a2e')
    sfill = PatternFill('solid', start_color='FCE8E6')
    wfill = PatternFill('solid', start_color='FFF8E1')
    fills = [PatternFill('solid', start_color='FFFFFF'),
             PatternFill('solid', start_color='EEF2FA')]
    ca  = Alignment(horizontal='center', vertical='center', wrap_text=True)
    la  = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    thin = Side(style='thin', color='CCCCCC')
    bd   = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ['No','승인일자','시간','병원','품목','HCP','진료과','가맹점','주소','승인금액','선정인원','비고']
    widths  = [5,   13,      8,    24,    14,    12,    12,    22,    26,    12,      9,       32]

    for ci,(h,w) in enumerate(zip(headers,widths),1):
        cell = ws.cell(row=1,column=ci,value=h)
        cell.font=hf; cell.fill=hfill; cell.alignment=ca; cell.border=bd
        ws.column_dimensions[get_column_letter(ci)].width=w

    tog=0; prev=None; rn=2
    for r in results:
        if r['No']!=prev: tog=(tog+1)%2; prev=r['No']
        is_sign = r['HCP']=='참석자 명단 서명 필요'
        fill = sfill if is_sign else (wfill if r['비고'] else fills[tog])

        vals = [r['No'],r['승인일자'],r['시간'],r['병원'],r['품목'],
                r['HCP'],r['진료과'],r['가맹점'],r['주소'],
                r['승인금액'],r['선정인원'],r['비고']]

        for ci,val in enumerate(vals,1):
            cell = ws.cell(row=rn,column=ci,value=val)
            cell.fill=fill; cell.border=bd
            cell.font=Font(name='Malgun Gothic', size=10,
                           bold=is_sign and ci==6,
                           color='C0392B' if (is_sign and ci==6) else '000000')
            cell.alignment = ca if ci in [1,2,3,10,11] else la
            if ci==10:
                try: cell.number_format='#,##0'
                except: pass
        rn+=1

    ws.freeze_panes='A2'
    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf

# ─── 영수증 합본 이미지 ───────────────────────────────────
def make_collage(images, cols=5):
    THUMB_W=500; GAP=8; BG=(235,237,242)
    rows=math.ceil(len(images)/cols)
    imgs=[]
    for img in images:
        r=THUMB_W/img.width
        imgs.append(img.resize((THUMB_W,int(img.height*r)),Image.LANCZOS))
    row_h=[max(imgs[r*cols:(r+1)*cols],key=lambda x:x.height).height for r in range(rows)]
    W=cols*THUMB_W+(cols+1)*GAP; H=sum(row_h)+(rows+1)*GAP
    canvas=Image.new('RGB',(W,H),BG)
    for idx,img in enumerate(imgs):
        row=idx//cols; col=idx%cols
        x=GAP+(cols-1-col)*(THUMB_W+GAP)
        y=GAP+sum(row_h[:row])+row*GAP
        canvas.paste(img,(x,y))
    return canvas

def date_label(date_str):
    try:
        d = datetime.strptime(date_str, '%Y-%m-%d')
        return f"{d.month}.{d.day}"
    except:
        return date_str

# ─── 메인 실행 ───────────────────────────────────────────
if st.button("🚀 매칭 시작", type="primary", use_container_width=True,
             disabled=not(api_key and receipt_files and excel_file)):

    st.divider()
    st.subheader("📋 처리 결과")

    with st.spinner("방문기록 엑셀 읽는 중..."):
        visits = load_visit_records(excel_file)
        st.success(f"✅ 방문기록 {len(visits)}건 로드 완료")

    st.write("**영수증 분석 중...**")
    receipt_data = []
    sorted_files = sorted(receipt_files, key=lambda f: f.name)
    prog = st.progress(0)

    for i, f in enumerate(sorted_files):
        with st.spinner(f"분석 중: {f.name}"):
            try:
                if i > 0:
                    time.sleep(4)
                f.seek(0)
                info = analyze_receipt_with_gemini(f)
                info['no'] = i + 1
                info['original_filename'] = f.name
                info['persons'] = calc_persons(info.get('amount', 0))
                info['need_sign'] = has_item_over_10000(info)
                sign_msg = " 🔴 서명필요" if info['need_sign'] else ""
                receipt_data.append(info)
                st.write(f"  ✅ {f.name} → {info.get('shop_name','')} / {info.get('amount',0):,}원 / {info.get('persons',0)}명{sign_msg}")
            except Exception as e:
                st.warning(f"  ⚠ {f.name} 분석 실패: {e}")
        prog.progress((i+1)/len(sorted_files))

    receipt_data.sort(key=lambda x: (x.get('date',''), x.get('time','')))
    for i, r in enumerate(receipt_data):
        r['no'] = i + 1

    with st.spinner("HCP 매칭 중..."):
        results = match_hcp(receipt_data, visits, min_days, max_per_month)

    sign_count = sum(1 for r in results if r['HCP'] == '참석자 명단 서명 필요')
    st.success(f"✅ 총 {len(results)}건 처리 완료  |  서명필요: {sign_count}건")
    st.dataframe(pd.DataFrame(results), use_container_width=True)

    # ─── ZIP 생성 ─────────────────────────────────────────
    st.write("**📦 ZIP 파일 생성 중...**")

    file_map = {f.name: f for f in sorted_files}
    sorted_rd = sorted(receipt_data, key=lambda x: (x.get('date',''), x.get('time','')))

    pil_images = []
    renamed = []

    for r in sorted_rd:
        f = file_map.get(r.get('original_filename',''))
        if f:
            f.seek(0)
            img = Image.open(f).convert('RGB')
            pil_images.append(img)
            time_label = r.get('time','').replace(':','.')
            new_name = f"{date_label(r.get('date',''))} {time_label} 영수증.jpg"
            renamed.append((new_name, img))

    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        # 1) 매칭 결과 엑셀
        excel_buf = make_excel(results)
        zf.writestr("HCP_매칭결과.xlsx", excel_buf.read())

        # 2) 개별 영수증
        for new_name, img in renamed:
            ibuf = io.BytesIO()
            img.save(ibuf, 'JPEG', quality=90)
            zf.writestr(f"영수증/{new_name}", ibuf.getvalue())

        # 3) 합본 이미지 (10장씩)
        for batch_idx in range(0, len(pil_images), 10):
            batch_imgs = pil_images[batch_idx:batch_idx+10]
            batch_rd   = sorted_rd[batch_idx:batch_idx+10]
            dates = [r.get('date','') for r in batch_rd if r.get('date','')]
            cname = f"{date_label(max(dates))} 증빙건.png" if dates else f"증빙건_{batch_idx//10+1}.png"
            collage = make_collage(batch_imgs)
            cbuf = io.BytesIO()
            collage.save(cbuf, 'PNG')
            zf.writestr(f"합본/{cname}", cbuf.getvalue())

    zip_buf.seek(0)
    all_dates = [r.get('date','') for r in sorted_rd if r.get('date','')]
    zip_name = f"{date_label(max(all_dates))} 증빙자료.zip" if all_dates else "증빙자료.zip"

    st.success("✅ ZIP 생성 완료!")
    st.download_button(
        label=f"📥 {zip_name} 다운로드",
        data=zip_buf,
        file_name=zip_name,
        mime="application/zip",
        use_container_width=True,
        type="primary"
    )
